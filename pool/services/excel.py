"""Servicio de generación y envío del Excel de predicciones.

El libro lleva una hoja por **ventana** de la quiniela: las de grupos
abren con una columna "Grupo"; las de eliminatoria, con "Partido" (el
``of_number``) y cierran con una columna "Avanza" (equipo que el jugador
cree que pasa por penales en un empate). Las banderas van como emoji junto
al nombre del equipo (en Windows pueden verse como las siglas del país).
"""

from io import BytesIO

from django.core.mail import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pool.models import Quiniela, Window
from pool.views.stages import render_window_sections

XLSX_MIMETYPE = (
    "application/vnd.openxmlformats-officedocument"
    ".spreadsheetml.sheet"
)

# Columnas comunes tras el identificador (Grupo/Partido).
_REST_HEADER = [
    "Local",
    "Goles local",
    "Goles visitante",
    "Visitante",
    "Día",
    "Hora",
    "Sede",
]

# Ancho por columna (caracteres openpyxl), de A a H.
_WIDTHS = [10, 24, 11, 13, 24, 12, 8, 28]

# Columnas a centrar (identificador, goles, día, hora): A, C, D, F, G.
_CENTERED = {"A", "C", "D", "F", "G"}

_FALLBACK_FILL = "D9D9D9"


def _team_cell(team, placeholder: str) -> str:
    """Nombre del equipo con su bandera emoji, o el placeholder textual."""
    if team is None:
        return placeholder
    flag = f"{team.flag_icon} " if team.flag_icon else ""
    return f"{flag}{team.name_es}"


def _advancing_cell(match) -> str:
    """Equipo que el jugador cree que pasa por penales (solo eliminatoria
    con empate pronosticado); vacío si no eligió. El servidor solo guarda
    ``advancing_team`` en empates válidos, así que basta mirar el id."""
    adv_id = getattr(match, "predicted_advancing_id", None)
    if not adv_id:
        return ""
    if match.home_team and match.home_team.id == adv_id:
        return _team_cell(match.home_team, "")
    if match.away_team and match.away_team.id == adv_id:
        return _team_cell(match.away_team, "")
    return ""


def _header_fill(color: str) -> PatternFill:
    """Relleno del encabezado tomando el color hex de la ventana."""
    hex_color = (color or "").lstrip("#")
    if len(hex_color) != 6:
        hex_color = _FALLBACK_FILL
    return PatternFill("solid", fgColor=f"FF{hex_color}")


def _style_sheet(sheet: Worksheet, is_group: bool, color: str) -> None:
    """Aplica anchos, encabezado coloreado, centrado y panel fijo.

    Las hojas de eliminatoria llevan una columna extra "Avanza" (I); el
    zip se corta solo en las de grupos (8 columnas)."""
    widths = _WIDTHS if is_group else [*_WIDTHS, 22]
    for letter, width in zip("ABCDEFGHI", widths):
        sheet.column_dimensions[letter].width = width

    bold = Font(bold=True)
    fill = _header_fill(color)
    center = Alignment(horizontal="center", vertical="center")
    for cell in sheet[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = center

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in _CENTERED:
                cell.alignment = center
    sheet.freeze_panes = "A2"


def _write_sheet(
    sheet: Worksheet, user, quiniela: Quiniela, window: Window
) -> None:
    """Vuelca las predicciones de una ventana en su hoja.

    El identificador de cada fila es la letra del grupo (ventana de grupos)
    o el ``of_number`` del partido (eliminatoria). El marcador queda en
    blanco si el usuario aún no predijo (o el cruce sigue siendo placeholder).
    """
    stages = list(window.stages.all())
    is_group = all(s.is_group for s in stages)
    header = ["Grupo" if is_group else "Partido", *_REST_HEADER]
    if not is_group:
        header.append("Avanza")
    sheet.append(header)

    for section in render_window_sections(user, quiniela, stages, is_group):
        for match in section["matches"]:
            ident = section["key"] if is_group else match.of_number
            row = [
                ident,
                _team_cell(match.home_team, match.home_placeholder),
                getattr(match, "predicted_home", ""),
                getattr(match, "predicted_away", ""),
                _team_cell(match.away_team, match.away_placeholder),
                match.local_day,
                match.local_time,
                match.stadium.name,
            ]
            if not is_group:
                row.append(_advancing_cell(match))
            sheet.append(row)
    _style_sheet(sheet, is_group, window.resolved_color())


def generate_excel(user, window: Window, quiniela: Quiniela) -> None:
    """Genera el libro completo de predicciones y lo envía por correo.

    Construye un ``Workbook`` con una hoja por **ventana** de la quiniela.
    ``window`` es la recién enviada y solo alimenta el texto del correo; el
    archivo refleja todo lo predicho hasta ahora en la quiniela. Los
    títulos de hoja se recortan a 31 caracteres (límite de openpyxl).
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    for w in quiniela.windows.prefetch_related("stages").order_by("order"):
        sheet = workbook.create_sheet(title=w.resolved_short_name()[:31])
        _write_sheet(sheet, user, quiniela, w)

    buffer = BytesIO()
    workbook.save(buffer)

    name = window.resolved_name()
    subject = f"Tus predicciones · {name}"
    body = (
        f"Acabas de enviar tus predicciones de {name}. Te adjunto el Excel "
        "con tu quiniela completa hasta ahora. Saludos!"
    )
    message = EmailMessage(subject=subject, body=body, to=[user.email])
    message.attach("predicciones.xlsx", buffer.getvalue(), XLSX_MIMETYPE)
    message.send()
