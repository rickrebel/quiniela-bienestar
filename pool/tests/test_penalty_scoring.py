"""Tests del scoring de la regla PENALTY (``Prediction.advancing_team``).

Cubre la evaluación (``recompute_all`` congela el bono PENALTY cuando el
jugador pronostica empate en eliminatoria y acierta al equipo que avanza por
penales) y la captura del pick vía el endpoint de autoguardado por partido.
"""

import json
from datetime import timedelta
from decimal import Decimal
from io import StringIO

from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook

from pool.models import (
    Prediction, Quiniela, User, UserQuiniela, Window)
from pool.services.evaluation import recompute_all
from pool.services.excel import _write_sheet
from tournament.models import Match, Stadium, Stage, Team


class PenaltyScoringTests(TestCase):
    """El bono PENALTY se congela solo en el caso correcto."""

    @classmethod
    def setUpTestData(cls):
        call_command("load_rules", stdout=StringIO())  # default = bienestar
        now = timezone.now()
        cls.group_stage = Stage.objects.create(
            key="GROUP_STAGE", name="Fase de grupos", short_name="grupos",
            color="#4CAF50", order=1, is_group=True,
            send_deadline=now - timedelta(days=1),
        )
        cls.knockout = Stage.objects.create(
            key="LAST_32", name="Dieciseisavos", short_name="16avos",
            color="#2196F3", order=2, send_deadline=now - timedelta(days=1),
        )
        cls.stadium = Stadium.objects.create(
            name="Estadio Azteca", city="CDMX", country="mx", utc_offset=-6,
        )
        cls.team_a = Team.objects.create(
            name="Mexico", name_es="México", fifa_code="MEX", group_name="A",
            confederation="concacaf",
        )
        cls.team_b = Team.objects.create(
            name="Canada", name_es="Canadá", fifa_code="CAN", group_name="A",
            confederation="concacaf",
        )
        # Knockout decidido por penales: empate a 120', gana A la tanda.
        cls.penalty_match = Match.objects.create(
            datetime=now - timedelta(hours=3), stage=cls.knockout,
            stadium=cls.stadium, home_team=cls.team_a, away_team=cls.team_b,
            of_number=73, status="FINISHED", home_goals=1, away_goals=1,
            decided_by=Match.PENALTY_SHOOTOUT,
            home_penalties=4, away_penalties=2,
        )
        # Knockout en tiempo regular (sin penales): A gana 2-1.
        cls.regular_match = Match.objects.create(
            datetime=now - timedelta(hours=3), stage=cls.knockout,
            stadium=cls.stadium, home_team=cls.team_a, away_team=cls.team_b,
            of_number=74, status="FINISHED", home_goals=2, away_goals=1,
            decided_by=Match.REGULAR,
        )
        # Empate de grupo (la regla PENALTY no aplica fuera de eliminatoria).
        cls.group_match = Match.objects.create(
            datetime=now - timedelta(hours=3), stage=cls.group_stage,
            stadium=cls.stadium, home_team=cls.team_a, away_team=cls.team_b,
            of_number=1, status="FINISHED", home_goals=1, away_goals=1,
            decided_by=Match.REGULAR,
        )

    def _predict(self, match, home, away, advancing=None, quiniela=None):
        quiniela = quiniela or Quiniela.objects.get(slug="bienestar")
        user = User.objects.create_user(
            f"u{Prediction.objects.count()}@x.com", first_name="U",
            is_active=True,
        )
        return Prediction.objects.create(
            user=user, quiniela=quiniela, match=match,
            home_goals=home, away_goals=away,
            advancing_team=advancing, date=timezone.now(),
        )

    def _eval(self, pred):
        """Reevalúa todo y devuelve (puntos, conjunto de códigos) frescos."""
        recompute_all()
        pred.refresh_from_db()
        codes = {r.code for r in pred.rules.all()}
        return pred.points, codes

    def test_exact_draw_correct_advancing_scores_penalty(self):
        pred = self._predict(self.penalty_match, 1, 1, advancing=self.team_a)
        points, codes = self._eval(pred)
        self.assertEqual(codes, {"RESULT", "EXACT", "PENALTY"})
        self.assertEqual(points, Decimal("5"))  # 3 + 1 + 1

    def test_inexact_draw_correct_advancing_scores_penalty(self):
        pred = self._predict(self.penalty_match, 2, 2, advancing=self.team_a)
        points, codes = self._eval(pred)
        self.assertEqual(codes, {"RESULT", "PENALTY"})
        self.assertEqual(points, Decimal("4"))  # 3 + 1

    def test_wrong_advancing_no_penalty(self):
        pred = self._predict(self.penalty_match, 1, 1, advancing=self.team_b)
        points, codes = self._eval(pred)
        self.assertEqual(codes, {"RESULT", "EXACT"})
        self.assertEqual(points, Decimal("4"))

    def test_no_advancing_no_penalty(self):
        pred = self._predict(self.penalty_match, 1, 1, advancing=None)
        _, codes = self._eval(pred)
        self.assertNotIn("PENALTY", codes)

    def test_non_draw_prediction_no_penalty(self):
        # Pronosticó 2-1 (no empate) en un partido que fue a penales: el
        # resultado es empate, así que ni acierta ni hay penal.
        pred = self._predict(self.penalty_match, 2, 1, advancing=self.team_a)
        points, codes = self._eval(pred)
        self.assertEqual(codes, set())
        self.assertEqual(points, Decimal("0"))

    def test_regular_time_draw_no_penalty(self):
        # advancing acierta a A, pero el partido se decidió en regular.
        pred = self._predict(self.regular_match, 2, 1, advancing=self.team_a)
        _, codes = self._eval(pred)
        self.assertNotIn("PENALTY", codes)

    def test_group_stage_draw_no_penalty(self):
        pred = self._predict(self.group_match, 1, 1, advancing=self.team_a)
        _, codes = self._eval(pred)
        self.assertNotIn("PENALTY", codes)

    def test_sanginiela_quiniela_without_penalty_rule(self):
        # Una predicción en sanginiela (sin la regla PENALTY) no puntúa el
        # penal aunque el pick sea correcto: el scoring es por quiniela.
        sanginiela = Quiniela.objects.get(slug="sanginiela")
        pred = self._predict(
            self.penalty_match, 1, 1, advancing=self.team_a,
            quiniela=sanginiela)
        points, codes = self._eval(pred)
        self.assertEqual(codes, {"RESULT", "EXACT"})
        self.assertEqual(points, Decimal("4"))


class AdvancingCaptureTests(TestCase):
    """El endpoint de autoguardado persiste/limpia ``advancing_team``."""

    @classmethod
    def setUpTestData(cls):
        now = timezone.now()
        cls.knockout = Stage.objects.create(
            key="LAST_32", name="Dieciseisavos", short_name="16avos",
            color="#2196F3", order=2,
            opens_at=now - timedelta(days=1),
            send_deadline=now + timedelta(days=1),
        )
        cls.group_stage = Stage.objects.create(
            key="GROUP_STAGE", name="Fase de grupos", short_name="grupos",
            color="#4CAF50", order=1, is_group=True,
            opens_at=now - timedelta(days=1),
            send_deadline=now + timedelta(days=1),
        )
        # Ventanas 1:1 (heredan calendario de su fase) + membresía: la señal
        # de UserQuiniela materializa los WindowUser editables.
        cls.quiniela = Quiniela.objects.create(name="Q", slug="q")
        cls.w_groups = Window.objects.create(quiniela=cls.quiniela, order=1)
        cls.w_groups.stages.add(cls.group_stage)
        cls.w_ko = Window.objects.create(quiniela=cls.quiniela, order=2)
        cls.w_ko.stages.add(cls.knockout)
        cls.stadium = Stadium.objects.create(
            name="Estadio Azteca", city="CDMX", country="mx", utc_offset=-6,
        )
        cls.team_a = Team.objects.create(
            name="Mexico", name_es="México", fifa_code="MEX", group_name="A",
            confederation="concacaf",
        )
        cls.team_b = Team.objects.create(
            name="Canada", name_es="Canadá", fifa_code="CAN", group_name="A",
            confederation="concacaf",
        )
        cls.outsider = Team.objects.create(
            name="Brazil", name_es="Brasil", fifa_code="BRA", group_name="B",
            confederation="conmebol",
        )
        cls.knockout_match = Match.objects.create(
            datetime=now + timedelta(days=2), stage=cls.knockout,
            stadium=cls.stadium, home_team=cls.team_a, away_team=cls.team_b,
            of_number=73,
        )
        cls.group_match = Match.objects.create(
            datetime=now + timedelta(days=2), stage=cls.group_stage,
            stadium=cls.stadium, home_team=cls.team_a, away_team=cls.team_b,
            of_number=1,
        )
        cls.player = User.objects.create_user(
            "ana@x.com", first_name="Ana", is_active=True,
        )
        # La señal de UserQuiniela crea sus WindowUser (ventanas abiertas y
        # sin vencer ⇒ editables).
        UserQuiniela.objects.create(user=cls.player, quiniela=cls.quiniela)

    def setUp(self):
        self.client.force_login(self.player)

    def _save(self, match, payload):
        return self.client.post(
            f"/q/prediction/{match.id}/",
            json.dumps(payload), content_type="application/json",
        )

    def _pred(self, match):
        return Prediction.objects.get(user=self.player, match=match)

    def test_draw_stores_advancing(self):
        self._save(self.knockout_match, {
            "home_goals": 1, "away_goals": 1,
            "advancing_team_id": self.team_b.id})
        self.assertEqual(self._pred(self.knockout_match).advancing_team_id,
                         self.team_b.id)

    def test_non_draw_clears_advancing(self):
        self._save(self.knockout_match, {
            "home_goals": 2, "away_goals": 1,
            "advancing_team_id": self.team_a.id})
        self.assertIsNone(self._pred(self.knockout_match).advancing_team_id)

    def test_foreign_team_ignored(self):
        self._save(self.knockout_match, {
            "home_goals": 1, "away_goals": 1,
            "advancing_team_id": self.outsider.id})
        self.assertIsNone(self._pred(self.knockout_match).advancing_team_id)

    def test_group_draw_does_not_store_advancing(self):
        self._save(self.group_match, {
            "home_goals": 1, "away_goals": 1,
            "advancing_team_id": self.team_a.id})
        self.assertIsNone(self._pred(self.group_match).advancing_team_id)

    def test_excel_knockout_sheet_has_advancing(self):
        Prediction.objects.create(
            user=self.player, quiniela=self.quiniela,
            match=self.knockout_match,
            home_goals=1, away_goals=1, advancing_team=self.team_b,
            date=timezone.now())
        sheet = Workbook().active
        _write_sheet(sheet, self.player, self.quiniela, self.w_ko)
        self.assertEqual(sheet.cell(row=1, column=9).value, "Avanza")
        row = next(r for r in sheet.iter_rows(min_row=2)
                   if r[0].value == self.knockout_match.of_number)
        self.assertEqual(row[8].value, "Canadá")  # sin flag_icon en fixture

    def test_excel_group_sheet_has_no_advancing_column(self):
        sheet = Workbook().active
        _write_sheet(sheet, self.player, self.quiniela, self.w_groups)
        self.assertEqual(sheet.cell(row=1, column=8).value, "Sede")
        self.assertIsNone(sheet.cell(row=1, column=9).value)
